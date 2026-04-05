import io
import json
import csv
from uuid import UUID
from datetime import datetime, timezone
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from database import get_db, SessionLocal
from models import (
    Person, Project, CaptureItem, CaptureItemPerson, CaptureItemProject,
    PersonProject, MeetingSession, ProfileLog, ItemNote, Setting, AIJob,
    ReportingLevel, ItemStatus
)

router = APIRouter(prefix="/api/import-export", tags=["import-export"])

REPORTING_LEVEL_MAP = {
    "executive": ReportingLevel.executive,
    "director": ReportingLevel.executive,
    "manager": ReportingLevel.manager,
    "employee": ReportingLevel.ic,
    "ic": ReportingLevel.ic,
    "peer": ReportingLevel.ic,
    "other": ReportingLevel.ic,
}

# ── Column name mappings for org XLSX import ──
ORG_COL_MAP = {
    "unique identifier": "external_id",
    "name": "name",
    "reports to": "reports_to",
    "line detail 1": "role",
    "line detail 2": "location",
    "line detail 3": "_unused",
    "organization name": "org_name",
}


# ── Import ──

@router.post("/preview")
async def preview_import(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Parse uploaded CSV/XLSX and return preview of changes."""
    content = await file.read()
    rows = _parse_file(file.filename, content)

    if not rows:
        raise HTTPException(400, "No data found in file")

    # Validate and categorize
    existing = {p.display_name.lower(): p for p in db.query(Person).all()}
    preview = []
    errors = []

    for i, row in enumerate(rows, start=2):  # Row 2+ (1 is header)
        display_name = (row.get("display_name") or row.get("display name") or "").strip()
        full_name = (row.get("full_name") or row.get("full name") or row.get("name") or "").strip()

        if not display_name and not full_name:
            errors.append({"row": i, "error": "Missing name and display name"})
            continue

        if not display_name:
            display_name = full_name
        if not full_name:
            full_name = display_name

        action = "update" if display_name.lower() in existing else "create"
        preview.append({
            "row": i,
            "action": action,
            "display_name": display_name,
            "full_name": full_name,
            "role": (row.get("title") or row.get("role") or "").strip(),
            "reporting_level": (row.get("reporting_level") or row.get("level") or row.get("director/manager/employee") or "").strip().lower(),
            "reporting_manager": (row.get("reporting_manager") or row.get("reporting manager") or row.get("manager") or "").strip(),
            "location": (row.get("location") or "").strip(),
            "address": (row.get("address") or "").strip(),
            "email": (row.get("email") or "").strip(),
        })

    # Check manager references
    all_names = {p["display_name"].lower() for p in preview}
    all_names.update(existing.keys())
    for p in preview:
        mgr = p["reporting_manager"]
        if mgr and mgr.lower() not in all_names:
            errors.append({"row": p["row"], "error": f"Reporting manager '{mgr}' not found"})

    creates = len([p for p in preview if p["action"] == "create"])
    updates = len([p for p in preview if p["action"] == "update"])

    return {"preview": preview, "errors": errors, "creates": creates, "updates": updates, "total": len(preview)}


@router.post("/commit")
async def commit_import(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Actually import the file — creates and updates people."""
    content = await file.read()
    rows = _parse_file(file.filename, content)

    if not rows:
        raise HTTPException(400, "No data found in file")

    existing = {p.display_name.lower(): p for p in db.query(Person).all()}
    created = 0
    updated = 0
    errors = []

    # First pass: create/update people (without manager links)
    for i, row in enumerate(rows, start=2):
        display_name = (row.get("display_name") or row.get("display name") or "").strip()
        full_name = (row.get("full_name") or row.get("full name") or row.get("name") or "").strip()
        if not display_name and not full_name:
            continue
        if not display_name:
            display_name = full_name
        if not full_name:
            full_name = display_name

        role = (row.get("title") or row.get("role") or "").strip()
        level_str = (row.get("reporting_level") or row.get("level") or row.get("director/manager/employee") or "").strip().lower()
        level = REPORTING_LEVEL_MAP.get(level_str, ReportingLevel.other)
        location = (row.get("location") or "").strip()
        address = (row.get("address") or "").strip()
        email = (row.get("email") or "").strip()

        person = existing.get(display_name.lower())
        if person:
            # Update
            person.name = full_name
            person.role = role or person.role
            person.reporting_level = level
            if email:
                person.email = email
            profile = {**({} if not person.profile else person.profile)}
            if location:
                profile["location"] = location
            if address:
                profile["address"] = address
            person.profile = profile
            from sqlalchemy.orm.attributes import flag_modified
            flag_modified(person, "profile")
            updated += 1
        else:
            # Create
            profile = {
                "spouse": "", "anniversary": "", "children": [],
                "pets": [], "birthday": "", "hobbies": "",
                "location": location, "address": address, "general": ""
            }
            person = Person(
                name=full_name, display_name=display_name, role=role,
                reporting_level=level, email=email, profile=profile,
            )
            db.add(person)
            db.flush()
            existing[display_name.lower()] = person
            created += 1

    # Second pass: resolve manager references
    for row in rows:
        display_name = (row.get("display_name") or row.get("display name") or "").strip()
        if not display_name:
            display_name = (row.get("full_name") or row.get("full name") or row.get("name") or "").strip()
        mgr_name = (row.get("reporting_manager") or row.get("reporting manager") or row.get("manager") or "").strip()

        if display_name and mgr_name:
            person = existing.get(display_name.lower())
            manager = existing.get(mgr_name.lower())
            if person and manager:
                person.manager_id = manager.id
            elif mgr_name:
                errors.append(f"Manager '{mgr_name}' not found for '{display_name}'")

    db.commit()
    return {"created": created, "updated": updated, "errors": errors}


def _find_import_roots(org_rows: list) -> set:
    """Find root external_ids of the import — people whose reports_to is not in the file."""
    all_ext_ids = {r["external_id"] for r in org_rows}
    roots = set()
    for r in org_rows:
        reports_to = r.get("reports_to", "").strip()
        if not reports_to or reports_to not in all_ext_ids:
            roots.add(r["external_id"])
    return roots


def _get_subtree_ext_ids(db, root_ext_ids: set) -> set:
    """Get all external_ids currently under the given root(s) in the DB. BFS tree walk."""
    if not root_ext_ids:
        return set()
    # Load all org-imported people with external_ids
    all_people = db.query(Person).filter(Person.external_id != None, Person.import_source == "org_import").all()
    ext_to_id = {p.external_id: p.id for p in all_people}
    id_to_ext = {p.id: p.external_id for p in all_people}
    children_map = {}
    for p in all_people:
        if p.manager_id:
            children_map.setdefault(p.manager_id, []).append(p.id)

    # Find the Person IDs for our roots
    root_person_ids = set()
    for ext_id in root_ext_ids:
        pid = ext_to_id.get(ext_id)
        if pid:
            root_person_ids.add(pid)

    # BFS from roots
    result_ext_ids = set()
    queue = list(root_person_ids)
    seen = set()
    while queue:
        current = queue.pop(0)
        if current in seen:
            continue
        seen.add(current)
        ext = id_to_ext.get(current)
        if ext:
            result_ext_ids.add(ext)
        queue.extend(children_map.get(current, []))
    return result_ext_ids


def _generate_unique_display_name(full_name: str, taken_names: set) -> str:
    """Generate a unique display name from a full name.
    Uses the full name as-is. If duplicate, appends a number: John Smith → John Smith 2."""
    candidate = full_name.strip()
    if candidate.lower() not in taken_names:
        return candidate
    for i in range(2, 100):
        c = f"{candidate} {i}"
        if c.lower() not in taken_names:
            return c
    return candidate


def _parse_org_xlsx(rows: list) -> tuple:
    """Normalize org XLSX rows. Returns (primary_rows, alias_map).
    alias_map maps duplicate/inherited ext_ids → primary ext_id for the same person.
    Deduplicates by name — first primary row wins, subsequent same-name rows become aliases."""
    all_rows = []
    inherited = []
    seen_ext_ids = set()
    for row in rows:
        mapped = {}
        for col_name, value in row.items():
            key = ORG_COL_MAP.get(col_name.strip().lower(), col_name.strip().lower())
            mapped[key] = (value or "").strip()
        if not mapped.get("external_id") or not mapped.get("name"):
            continue
        if mapped["external_id"] in seen_ext_ids:
            continue
        seen_ext_ids.add(mapped["external_id"])
        org_name = mapped.get("org_name", "").lower()
        if "(inherited)" in org_name:
            inherited.append(mapped)
        else:
            all_rows.append(mapped)

    # Deduplicate primary rows by name — first occurrence wins, rest become aliases
    primary = []
    primary_by_name = {}  # name_lower → primary ext_id
    alias_map = {}

    for r in all_rows:
        name_l = r["name"].lower().strip()
        if name_l in primary_by_name:
            # Same name already has a primary — alias this ext_id to the first one
            alias_map[r["external_id"]] = primary_by_name[name_l]
        else:
            primary.append(r)
            primary_by_name[name_l] = r["external_id"]

    # Also alias inherited rows to their primary
    for r in inherited:
        name_l = r["name"].lower().strip()
        primary_ext = primary_by_name.get(name_l)
        if primary_ext:
            alias_map[r["external_id"]] = primary_ext

    return primary, alias_map


def _resolve_file_managers(org_rows, alias_map):
    """Build a map of file ext_id → manager's name (resolved within the file).
    Used for name+manager composite matching."""
    ext_to_name = {r["external_id"]: r["name"].strip().lower() for r in org_rows}
    # Also add alias targets
    for alias_id, primary_id in alias_map.items():
        if primary_id in ext_to_name:
            ext_to_name[alias_id] = ext_to_name[primary_id]

    result = {}  # ext_id → manager_name (lowercase)
    for r in org_rows:
        reports_to = r.get("reports_to", "").strip()
        if reports_to:
            resolved = alias_map.get(reports_to, reports_to)
            result[r["external_id"]] = ext_to_name.get(resolved, "")
        else:
            result[r["external_id"]] = ""
    return result


def _build_db_name_manager_map(db):
    """Build lookup of (name_lower, manager_name_lower) → Person for org-imported people."""
    people = db.query(Person).filter(Person.import_source == "org_import").all()
    manager_ids = {p.manager_id for p in people if p.manager_id}
    managers = {p.id: p.name.lower().strip() for p in db.query(Person).filter(Person.id.in_(manager_ids)).all()} if manager_ids else {}

    by_name_mgr = {}  # (name_lower, mgr_name_lower) → Person
    by_name = {}  # name_lower → [Person, ...]
    for p in people:
        name_l = p.name.lower().strip()
        mgr_name = managers.get(p.manager_id, "") if p.manager_id else ""
        by_name_mgr[(name_l, mgr_name)] = p
        by_name.setdefault(name_l, []).append(p)
    return by_name_mgr, by_name


def _get_subtree_names(db, root_names: set) -> set:
    """Get all names currently under the given root names in the DB. BFS tree walk."""
    if not root_names:
        return set()
    all_people = db.query(Person).filter(Person.import_source == "org_import", Person.is_archived == False).all()
    name_to_id = {}
    id_to_name = {}
    children_map = {}
    for p in all_people:
        name_to_id.setdefault(p.name.lower().strip(), []).append(p.id)
        id_to_name[p.id] = p.name.lower().strip()
        if p.manager_id:
            children_map.setdefault(p.manager_id, []).append(p.id)

    root_ids = set()
    for rn in root_names:
        root_ids.update(name_to_id.get(rn, []))

    result_names = set()
    queue = list(root_ids)
    seen = set()
    while queue:
        current = queue.pop(0)
        if current in seen:
            continue
        seen.add(current)
        n = id_to_name.get(current)
        if n:
            result_names.add(n)
        queue.extend(children_map.get(current, []))
    return result_names


@router.post("/org-preview")
async def org_preview(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Preview org chart XLSX import — name+manager matching with review flags."""
    content = await file.read()
    rows = _parse_file(file.filename, content)
    if not rows:
        raise HTTPException(400, "No data found in file")

    org_rows, alias_map = _parse_org_xlsx(rows)
    if not org_rows:
        raise HTTPException(400, "No valid rows found. Expected columns: Unique Identifier, Name, Reports To, etc.")

    # Resolve manager names within the file
    file_mgr_map = _resolve_file_managers(org_rows, alias_map)

    # Build DB lookup by (name, manager_name)
    by_name_mgr, by_name = _build_db_name_manager_map(db)

    creates = []
    updates = []
    unchanged = 0
    archives = []
    review = []  # ambiguous cases for user review
    file_names = set()  # track all names in file for archive detection

    for r in org_rows:
        ext_id = r["external_id"]
        name = r.get("name", "").strip()
        name_l = name.lower()
        role = r.get("role", "")
        location = r.get("location", "")
        org_name = r.get("org_name", "")
        is_leader = bool(org_name)
        file_mgr_name = file_mgr_map.get(ext_id, "")
        file_names.add(name_l)

        # Priority 1: exact name + manager match
        person = by_name_mgr.get((name_l, file_mgr_name))

        # Priority 1b: if file manager is blank (root of partial export), match by name only
        # without treating it as a manager change — their manager just isn't in this file
        if not person and not file_mgr_name and name_l in by_name:
            candidates = by_name[name_l]
            if len(candidates) == 1:
                person = candidates[0]

        if person:
            # Confident match — check for field changes
            changes = {}
            if role and role != (person.role or ""):
                changes["role"] = {"from": person.role or "", "to": role}
            profile = person.profile or {}
            if location and location != profile.get("location", ""):
                changes["location"] = {"from": profile.get("location", ""), "to": location}
            if changes:
                updates.append({"name": name, "display_name": person.display_name, "changes": changes, "match": "name+manager"})
            else:
                unchanged += 1
        elif name_l in by_name:
            # Name exists but manager is different
            existing = by_name[name_l]
            if len(existing) == 1 and not file_mgr_name:
                # Single match, file manager blank (partial export root) — safe match
                p = existing[0]
                changes = {}
                if role and role != (p.role or ""):
                    changes["role"] = {"from": p.role or "", "to": role}
                profile = p.profile or {}
                if location and location != profile.get("location", ""):
                    changes["location"] = {"from": profile.get("location", ""), "to": location}
                if changes:
                    updates.append({"name": name, "display_name": p.display_name, "changes": changes, "match": "name_only"})
                else:
                    unchanged += 1
            elif len(existing) == 1:
                # Single match but different manager — likely different person with same name
                p = existing[0]
                old_mgr = p.manager.name if p.manager_id and p.manager else ""
                review.append({
                    "name": name,
                    "display_name": p.display_name,
                    "old_manager": old_mgr,
                    "new_manager": file_mgr_name,
                    "reason": f"Same name exists under different manager — will create as new person",
                })
                creates.append({"name": name, "role": role, "location": location, "is_leader": is_leader})
            else:
                # Multiple people with this name — ambiguous
                review.append({
                    "name": name,
                    "display_name": None,
                    "old_manager": ", ".join(p.display_name for p in existing),
                    "new_manager": file_mgr_name,
                    "reason": f"Multiple '{name}' exist — cannot auto-match, will create new",
                })
                creates.append({"name": name, "role": role, "location": location, "is_leader": is_leader})
        else:
            creates.append({"name": name, "role": role, "location": location, "is_leader": is_leader})

    # Scope archive to the import's subtree
    root_names = set()
    for r in org_rows:
        reports_to = r.get("reports_to", "").strip()
        if not reports_to or alias_map.get(reports_to, reports_to) not in {rr["external_id"] for rr in org_rows}:
            root_names.add(r["name"].strip().lower())
    subtree_names = _get_subtree_names(db, root_names)

    for sn in subtree_names:
        if sn not in file_names:
            people = by_name.get(sn, [])
            for p in people:
                if not p.is_archived:
                    archives.append({"name": p.name, "display_name": p.display_name})

    return {
        "creates": creates,
        "updates": updates,
        "archives": archives,
        "review": review,
        "unchanged_count": unchanged,
        "total_rows": len(org_rows),
    }


@router.post("/org-commit")
async def org_commit(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Commit org chart XLSX import — name+manager matching."""
    content = await file.read()
    rows = _parse_file(file.filename, content)
    if not rows:
        raise HTTPException(400, "No data found in file")

    org_rows, alias_map = _parse_org_xlsx(rows)
    if not org_rows:
        raise HTTPException(400, "No valid org rows found")

    # Resolve manager names within the file
    file_mgr_map = _resolve_file_managers(org_rows, alias_map)

    # Build DB lookups
    by_name_mgr, by_name = _build_db_name_manager_map(db)

    all_people = db.query(Person).all()
    taken_display_names = {p.display_name.lower() for p in all_people}

    created_count = 0
    updated_count = 0
    archived_count = 0
    errors = []
    file_names = set()

    # Map file ext_id → Person (for manager resolution within file)
    ext_to_person = {}

    # Pass 1: Match and create/update people
    matched_person_ids = set()  # track which DB people have been matched to avoid double-matching
    for r in org_rows:
        ext_id = r["external_id"]
        name = r.get("name", "").strip()
        name_l = name.lower()
        role = r.get("role", "").strip()
        location = r.get("location", "").strip()
        file_mgr_name = file_mgr_map.get(ext_id, "")
        file_names.add(name_l)

        # Try name+manager match first (confident)
        person = by_name_mgr.get((name_l, file_mgr_name))
        if person and person.id in matched_person_ids:
            person = None  # already matched to a different file row

        # Fallback: name-only if file manager is blank (partial export root)
        if not person and not file_mgr_name and name_l in by_name:
            candidates = [c for c in by_name[name_l] if c.id not in matched_person_ids]
            if len(candidates) == 1:
                person = candidates[0]

        if person:
            # Update existing
            matched_person_ids.add(person.id)
            old_name = person.name
            if name and name != old_name:
                person.name = name
                if person.display_name == old_name:
                    taken_display_names.discard(person.display_name.lower())
                    new_dn = _generate_unique_display_name(name, taken_display_names)
                    person.display_name = new_dn
                    taken_display_names.add(new_dn.lower())
            if role:
                person.role = role
            if location:
                import copy
                profile = copy.deepcopy(person.profile or {})
                profile["location"] = location
                person.profile = profile
                from sqlalchemy.orm.attributes import flag_modified
                flag_modified(person, "profile")
            person.import_source = "org_import"
            person.external_id = ext_id
            if person.is_archived:
                person.is_archived = False
            ext_to_person[ext_id] = person
            updated_count += 1
        else:
            # No confident match — create new person
            display_name = _generate_unique_display_name(name, taken_display_names)
            taken_display_names.add(display_name.lower())
            profile = {
                "spouse": "", "anniversary": "", "children": [],
                "pets": [], "birthday": "", "hobbies": "",
                "location": location, "address": "", "general": ""
            }
            person = Person(
                name=name, display_name=display_name, role=role,
                reporting_level=ReportingLevel.ic, import_source="org_import",
                external_id=ext_id, profile=profile,
            )
            db.add(person)
            db.flush()
            ext_to_person[ext_id] = person
            created_count += 1

    # Pass 2: Resolve manager references using file ext_ids (with alias resolution)
    for r in org_rows:
        ext_id = r["external_id"]
        reports_to = r.get("reports_to", "").strip()
        person = ext_to_person.get(ext_id)
        if person and reports_to:
            resolved_to = alias_map.get(reports_to, reports_to)
            manager = ext_to_person.get(resolved_to)
            if manager:
                person.manager_id = manager.id
            # else: Reports To references someone not in this file — leave existing manager
        # elif person and not reports_to: blank Reports To = root of this export, keep existing manager

    # Pass 3: Archive departed — scoped to import subtree by name
    root_names = set()
    for r in org_rows:
        reports_to = r.get("reports_to", "").strip()
        if not reports_to or alias_map.get(reports_to, reports_to) not in {rr["external_id"] for rr in org_rows}:
            root_names.add(r["name"].strip().lower())
    subtree_names = _get_subtree_names(db, root_names)

    for sn in subtree_names:
        if sn not in file_names:
            people_list = by_name.get(sn, [])
            for p in people_list:
                if not p.is_archived:
                    p.is_archived = True
                    archived_count += 1

    # Pass 4: Infer reporting_level from title + tree structure
    import re as _re
    _exec_pattern = _re.compile(r'\b(evp|svp|vp|vice president|chief|president|head of)\b', _re.IGNORECASE)
    _director_pattern = _re.compile(r'\b(sr\.?\s*director|senior\s+director|director)\b', _re.IGNORECASE)
    has_reports = set()
    for person in ext_to_person.values():
        if person.manager_id:
            has_reports.add(person.manager_id)
    for person in ext_to_person.values():
        role_lower = (person.role or "")
        is_exec = bool(_exec_pattern.search(role_lower))
        is_director = bool(_director_pattern.search(role_lower))
        if is_exec:
            person.reporting_level = ReportingLevel.executive
        elif is_director:
            person.reporting_level = ReportingLevel.director
        elif person.id in has_reports:
            person.reporting_level = ReportingLevel.manager
        else:
            person.reporting_level = ReportingLevel.ic

    # Pass 5: Fix display names for imported people whose display_name == name (not customized)
    taken_dn = set()
    for person in ext_to_person.values():
        if person.display_name != person.name:
            taken_dn.add(person.display_name.lower())
    # Also include manually-created people's display names
    for p in all_people:
        if not p.external_id:
            taken_dn.add(p.display_name.lower())
    # Now re-generate for those whose display_name == name
    for person in ext_to_person.values():
        if person.display_name == person.name:
            new_dn = _generate_unique_display_name(person.name, taken_dn)
            person.display_name = new_dn
            taken_dn.add(new_dn.lower())

    db.commit()
    return {
        "created": created_count,
        "updated": updated_count,
        "archived": archived_count,
        "errors": errors,
        "total": len(org_rows),
    }


def _parse_file(filename: str, content: bytes) -> list:
    """Parse CSV or XLSX file into list of dicts."""
    if filename.endswith(".xlsx"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                return []
            headers = [str(h).strip().lower() if h else "" for h in rows[0]]
            return [{headers[j]: (str(cell).strip() if cell else "") for j, cell in enumerate(row) if j < len(headers)} for row in rows[1:]]
        except ImportError:
            raise HTTPException(400, "Excel support not available. Please use CSV format.")
    else:
        # CSV
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        return [{k.strip().lower(): (v.strip() if v else "") for k, v in row.items()} for row in reader]


# ── Export: Team Directory ──

@router.get("/export/team")
def export_team(
    columns: str = Query("display_name,name,role,reporting_level,location,address,email,manager"),
    format: str = Query("csv"),
    db: Session = Depends(get_db),
):
    """Export team directory as CSV or XLSX."""
    cols = [c.strip() for c in columns.split(",")]
    people = db.query(Person).filter(Person.is_archived == False).order_by(Person.display_name).all()

    rows = []
    for p in people:
        profile = p.profile or {}
        manager_name = ""
        if p.manager_id and p.manager:
            manager_name = p.manager.display_name
        row = {}
        for col in cols:
            if col == "display_name":
                row["Display Name"] = p.display_name
            elif col == "name":
                row["Full Name"] = p.name
            elif col == "role":
                row["Title"] = p.role or ""
            elif col == "reporting_level":
                row["Level"] = p.reporting_level.value if p.reporting_level else ""
            elif col == "location":
                row["Location"] = profile.get("location", "")
            elif col == "address":
                row["Address"] = profile.get("address", "")
            elif col == "email":
                row["Email"] = p.email or ""
            elif col == "manager":
                row["Reporting Manager"] = manager_name
            elif col == "spouse":
                row["Spouse"] = profile.get("spouse", "")
            elif col == "birthday":
                row["Birthday"] = profile.get("birthday", "")
            elif col == "children":
                row["Children"] = ", ".join(profile.get("children", []))
            elif col == "hobbies":
                row["Hobbies"] = profile.get("hobbies", "")
        rows.append(row)

    if format == "xlsx":
        return _export_xlsx(rows, "team_directory")
    else:
        return _export_csv(rows, "team_directory")


# ── Export: Full Backup ──

@router.get("/export/backup")
def export_backup(db: Session = Depends(get_db)):
    """Full JSON backup of all Ledger data."""
    from routers.captures import item_to_response

    people = db.query(Person).all()
    projects = db.query(Project).all()
    items = db.query(CaptureItem).order_by(CaptureItem.created_at).all()
    meetings = db.query(MeetingSession).all()
    logs = db.query(ProfileLog).all()
    settings = db.query(Setting).all()
    person_projects = db.query(PersonProject).all()

    backup = {
        "exported_at": datetime.now(timezone.utc).isoformat(),
        "version": "1.0",
        "people": [{
            "id": str(p.id), "name": p.name, "display_name": p.display_name,
            "role": p.role, "reporting_level": p.reporting_level.value if p.reporting_level else "other",
            "email": p.email, "is_archived": p.is_archived,
            "profile": p.profile, "avatar": p.avatar,
            "manager_id": str(p.manager_id) if p.manager_id else None,
            "created_at": p.created_at.isoformat() if p.created_at else None,
        } for p in people],
        "projects": [{
            "id": str(p.id), "name": p.name, "short_code": p.short_code,
            "status": p.status.value if p.status else "active",
            "is_archived": p.is_archived, "context_notes": p.context_notes,
            "created_at": p.created_at.isoformat() if p.created_at else None,
        } for p in projects],
        "items": [item_to_response(i) for i in items],
        "meetings": [{
            "id": str(m.id), "started_at": m.started_at.isoformat() if m.started_at else None,
            "ended_at": m.ended_at.isoformat() if m.ended_at else None,
            "person_id": str(m.person_id) if m.person_id else None,
            "project_id": str(m.project_id) if m.project_id else None,
            "items_resolved": m.items_resolved, "items_added": m.items_added,
            "ai_summary": m.ai_summary,
        } for m in meetings],
        "profile_logs": [{
            "id": str(l.id), "created_at": l.created_at.isoformat(),
            "log_type": l.log_type.value, "content": l.content,
            "person_id": str(l.person_id) if l.person_id else None,
            "project_id": str(l.project_id) if l.project_id else None,
        } for l in logs],
        "person_projects": [{
            "person_id": str(pp.person_id), "project_id": str(pp.project_id)
        } for pp in person_projects],
        "settings": {s.key: s.value for s in settings},
    }

    content = json.dumps(backup, indent=2, default=str)
    return StreamingResponse(
        io.BytesIO(content.encode()),
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename=ledger_backup_{datetime.now().strftime('%Y%m%d')}.json"}
    )


# ── Helpers ──

def _export_csv(rows: list, name: str):
    if not rows:
        return StreamingResponse(io.BytesIO(b"No data"), media_type="text/csv")
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=rows[0].keys())
    writer.writeheader()
    writer.writerows(rows)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={name}.csv"}
    )


def _export_xlsx(rows: list, name: str):
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        if rows:
            ws.append(list(rows[0].keys()))
            for row in rows:
                ws.append(list(row.values()))
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={name}.xlsx"}
        )
    except ImportError:
        raise HTTPException(400, "Excel export not available. Use CSV format.")
